import asyncio
import calendar
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pyodide.http import pyfetch
from pyodide.ffi import create_proxy
from pyscript import document
from js import Blob, URL, Uint8Array, document as js_document

# -------------------------------------------------------------------
# TRADUCCIÓN DE DÍAS A ESPAÑOL
# -------------------------------------------------------------------
TRAD_DIAS = {
    "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Mierc.", "Thursday": "Jueves", "Friday": "Viernes",
    "lunes": "Lunes", "martes": "Martes", "miércoles": "Mierc.", "miercoles": "Mierc.", "jueves": "Jueves", "viernes": "Viernes"
}

def normalizar_dia(d_str):
    d_clean = str(d_str).strip()
    return TRAD_DIAS.get(d_clean, TRAD_DIAS.get(d_clean.capitalize(), d_clean))

# -------------------------------------------------------------------
# CONSULTA DE FERIADOS API
# -------------------------------------------------------------------
async def consultar_feriados(anio, mes):
    url = f"https://nolaborables.com.ar/api/v2/feriados/{anio}"
    feriados = []
    try:
        response = await pyfetch(url, method="GET")
        if response.ok:
            data = await response.json()
            feriados = [int(f["dia"]) for f in data if int(f["mes"]) == mes and f.get("tipo") != "opcional"]
    except Exception as e:
        print(f"Aviso API feriados: {e}")
    return feriados

# -------------------------------------------------------------------
# DESCARGA DE EXCEL EN NAVEGADOR (CONVERSIÓN Uint8Array)
# -------------------------------------------------------------------
def descargar_excel(wb, filename):
    stream = BytesIO()
    wb.save(stream)
    raw_bytes = stream.getvalue()
    
    js_array = Uint8Array.new(len(raw_bytes))
    js_array.assign(raw_bytes)
    
    blob = Blob.new([js_array], {type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    url = URL.createObjectURL(blob)
    
    link = js_document.createElement("a")
    link.href = url
    link.download = filename
    link.click()
    URL.revokeObjectURL(url)

# -------------------------------------------------------------------
# PARSEO DE PERSONAL DINÁMICO DESDE EL DOM
# -------------------------------------------------------------------
def parse_personal_dinamico(personal_json_str):
    grupo_gis = set()
    duos_trios = set()
    resto = set()
    grupo_aho = []
    grupo_bho = []

    if personal_json_str.strip():
        try:
            items = json.loads(personal_json_str)
            for item in items:
                name = item.get("name")
                group = item.get("group")
                ho = item.get("ho")

                if not name: continue

                if group == "GIS": grupo_gis.add(name)
                elif group == "DUOS": duos_trios.add(name)
                else: resto.add(name)

                if ho == "AHO": grupo_aho.append(name)
                elif ho == "BHO": grupo_bho.append(name)
        except Exception as e:
            print(f"Error parseando personal: {e}")

    if not grupo_gis and not duos_trios and not resto:
        grupo_gis = {"Ledesma", "Itzi", "Fuentes", "Romero", "Coronel", "Nahuel", "Herrera"}
        duos_trios = {"Aranda", "Navarro", "Neri", "Tamara", "Alegre"}
        resto = {"Gonzalote", "Lezcano"}
        grupo_aho = ["Aranda", "Alegre", "Nahuel", "Fuentes", "Romero"]
        grupo_bho = ["Ledesma", "Lezcano", "Itzi", "Navarro", "Gonzalote"]

    personal_total = grupo_gis | duos_trios | resto
    return grupo_gis, duos_trios, resto, personal_total, grupo_aho, grupo_bho

# -------------------------------------------------------------------
# PARSEO DE NOVEDADES DESDE EL CALENDARIO INTERACTIVO (JSON)
# -------------------------------------------------------------------
def parse_novedades_json(json_str, anio, mes):
    conae_dict = {}
    licencias_dict = {} # emp: { date: detail_str }
    francos_dict = {}
    mananas_dict = {}
    tardes_dict = {}
    ho_excepcional_dict = {}
    feriados_manuales = []

    if not json_str.strip():
        return conae_dict, licencias_dict, francos_dict, mananas_dict, tardes_dict, ho_excepcional_dict, feriados_manuales

    try:
        items = json.loads(json_str)
        for item in items:
            emp = item.get("emp")
            date = item.get("date")
            ntype = item.get("type")
            detail = item.get("detail", "").strip()

            if not date or not ntype:
                continue

            if ntype == "FERIADO":
                feriados_manuales.append(date)
            elif ntype == "HO_EXCEPCIONAL" and emp:
                ho_excepcional_dict.setdefault(emp, []).append(date)
            elif ntype == "CONAE" and emp:
                conae_dict.setdefault(emp, []).append(date)
            elif ntype == "LICENCIA" and emp:
                licencias_dict.setdefault(emp, {})[date] = detail if detail else "Licencia"
            elif ntype == "FRANCO" and emp:
                francos_dict.setdefault(emp, []).append(date)
            elif ntype == "MANANA" and emp:
                mananas_dict.setdefault(emp, []).append(date)
            elif ntype == "TARDE" and emp:
                w_day_idx = calendar.weekday(anio, mes, date)
                if w_day_idx < 5:
                    day_names = ["Lunes", "Martes", "Mierc.", "Jueves", "Viernes"]
                    tardes_dict.setdefault(emp, []).append(day_names[w_day_idx])
    except Exception as e:
        print(f"Error parseando novedades JSON: {e}")

    return conae_dict, licencias_dict, francos_dict, mananas_dict, tardes_dict, ho_excepcional_dict, feriados_manuales

# -------------------------------------------------------------------
# MOTOR DE RESOLUCIÓN DE PLANILLA (SOLVER DINÁMICO)
# -------------------------------------------------------------------
def resolver_planilla(anio, mes, feriados_extra, conae_dict, licencias_dict, francos_solicitados_dict, mananas_solicitadas_dict, tardes_solicitadas_dict, ho_excepcional_dict, bloqueados_alfa, personal_info):
    grupo_gis, duos_trios, resto, personal_total, grupo_aho, grupo_bho = personal_info

    bloqueados_tarde = {"Herrera", "Itzi"}
    base_bloqueados_monitores = {"Itzi", "Alegre", "Nahuel", "Gonzalote", "Coronel", "Tamara", "Herrera"}
    bloqueados_monitores = base_bloqueados_monitores | set(bloqueados_alfa)
    
    elegibles_monitores = [p for p in personal_total if p not in bloqueados_monitores]
    if not elegibles_monitores:
        elegibles_monitores = [p for p in personal_total if p not in base_bloqueados_monitores]
        if not elegibles_monitores:
            elegibles_monitores = list(personal_total)

    cal = calendar.monthcalendar(anio, mes)
    weeks_data = []
    week_idx = 1
    for week in cal:
        business_days = [(d, calendar.day_name[i]) for i, d in enumerate(week[:5]) if d != 0]
        if business_days:
            weeks_data.append((week_idx, business_days))
            week_idx += 1

    schedule = {}
    early_counts = {p: 0 for p in elegibles_monitores}

    for w_num, days in weeks_data:
        dias_numeros = [d for d, _ in days]
        tiene_feriado = any(d in feriados_extra for d in dias_numeros)
        is_s1_s3 = (w_num % 2 != 0)

        for date, day_name_raw in days:
            day_name = normalizar_dia(day_name_raw)

            schedule[date] = {
                "day_name": day_name,
                "feriado": (date in feriados_extra),
                "ho": [],
                "licencia": {},
                "conae": [],
                "francos": [],
                "manana": [],
                "tarde": [],
                "early_person": None
            }

            if schedule[date]["feriado"]:
                continue

            # Home Office de grupo rotativo
            if day_name == "Lunes":
                if not is_s1_s3:
                    schedule[date]["ho"].extend(grupo_aho)
                else:
                    schedule[date]["ho"].extend(grupo_bho)
            elif day_name == "Viernes":
                if is_s1_s3:
                    schedule[date]["ho"].extend(grupo_aho)
                else:
                    schedule[date]["ho"].extend(grupo_bho)
            
            # HO Fijo Aranda los Martes
            if day_name == "Martes" and "Aranda" in personal_total and "Aranda" not in schedule[date]["ho"]:
                schedule[date]["ho"].append("Aranda")

            # Home Office Excepcional (entre semana fuera de grupos)
            for emp, h_days in ho_excepcional_dict.items():
                if date in h_days and emp not in schedule[date]["ho"]:
                    schedule[date]["ho"].append(emp)

            # Asignaciones a CONAE
            for emp, c_days in conae_dict.items():
                if date in c_days:
                    schedule[date]["conae"].append(emp)
                    if emp in schedule[date]["ho"]:
                        schedule[date]["ho"].remove(emp)

            # Licencias
            for emp, l_map in licencias_dict.items():
                if date in l_map:
                    schedule[date]["licencia"][emp] = l_map[date]
                    if emp in schedule[date]["ho"]:
                        schedule[date]["ho"].remove(emp)

            # Francos
            if not tiene_feriado:
                francos_hoy = []

                for emp, f_dias in francos_solicitados_dict.items():
                    if date in f_dias and emp not in schedule[date]["ho"] and emp not in schedule[date]["licencia"] and emp not in schedule[date]["conae"]:
                        francos_hoy.append(emp)

                ausentes_hoy = set(schedule[date]["ho"]) | set(schedule[date]["licencia"].keys()) | set(schedule[date]["conae"]) | set(francos_hoy)
                elegibles_franco = []
                for emp in (personal_total - ausentes_hoy - {"Coronel"}):
                    es_aho = emp in grupo_aho
                    es_bho = emp in grupo_bho

                    bloqueado_por_ho = False
                    if day_name == "Martes" and ((es_aho and not is_s1_s3) or (es_bho and is_s1_s3) or emp == "Aranda"):
                        bloqueado_por_ho = True
                    if day_name == "Jueves" and ((es_aho and is_s1_s3) or (es_bho and not is_s1_s3)):
                        bloqueado_por_ho = True
                    
                    if not bloqueado_por_ho:
                        elegibles_franco.append(emp)

                if day_name == "Martes":
                    francos_hoy.extend([e for e in ["Fuentes", "Lezcano", "Itzi"] if e in elegibles_franco and e not in francos_hoy][:2 - len(francos_hoy)])
                elif day_name == "Mierc.":
                    francos_hoy.extend([e for e in ["Nahuel", "Navarro", "Aranda"] if e in elegibles_franco and e not in francos_hoy][:2 - len(francos_hoy)])
                elif day_name == "Jueves":
                    francos_hoy.extend([e for e in ["Gonzalote", "Alegre", "Romero"] if e in elegibles_franco and e not in francos_hoy][:2 - len(francos_hoy)])
                elif day_name == "Viernes":
                    francos_hoy.extend([e for e in ["Ledesma", "Fuentes"] if e in elegibles_franco and e not in francos_hoy][:1 - len(francos_hoy)])

                schedule[date]["francos"] = francos_hoy

            disponibles = list(personal_total - set(schedule[date]["ho"]) - set(schedule[date]["licencia"].keys()) - set(schedule[date]["conae"]) - set(schedule[date]["francos"]))

            # Requerimiento forzado de Mañana
            must_manana = []
            for emp, m_dias in mananas_solicitadas_dict.items():
                if date in m_dias and emp in disponibles:
                    must_manana.append(emp)

            # Turno Tarde
            must_tarde = []
            for emp, t_dias in tardes_solicitadas_dict.items():
                if day_name in t_dias and emp in disponibles and emp not in bloqueados_tarde and emp not in must_manana:
                    must_tarde.append(emp)

            if day_name == "Lunes":
                if "Coronel" in disponibles and "Coronel" not in must_manana: must_tarde.append("Coronel")
                if "Aranda" in disponibles and "Aranda" not in must_manana: must_tarde.append("Aranda")
            elif day_name == "Martes":
                if "Alegre" in disponibles and "Alegre" not in must_manana: must_tarde.append("Alegre")

            tarde_assigned = list(set(must_tarde) - bloqueados_tarde - set(must_manana))
            gis_en_tarde = any(p in grupo_gis for p in tarde_assigned)

            restantes = [p for p in disponibles if p not in tarde_assigned and p not in bloqueados_tarde and p not in must_manana]
            
            if not gis_en_tarde:
                gis_disponibles = [p for p in restantes if p in grupo_gis]
                if gis_disponibles:
                    p_gis = gis_disponibles[0]
                    tarde_assigned.append(p_gis)
                    restantes.remove(p_gis)

            while len(tarde_assigned) < 2 and restantes:
                p = restantes.pop(0)
                tarde_assigned.append(p)

            manana_assigned = [p for p in disponibles if p not in tarde_assigned]

            schedule[date]["tarde"] = tarde_assigned
            schedule[date]["manana"] = manana_assigned

            posibles_monitor = [p for p in manana_assigned if p in elegibles_monitores]
            if posibles_monitor:
                posibles_monitor.sort(key=lambda x: early_counts.get(x, 0))
                elegido = posibles_monitor[0]
                early_counts[elegido] = early_counts.get(elegido, 0) + 1
                schedule[date]["early_person"] = elegido

    return weeks_data, schedule

# -------------------------------------------------------------------
# ARMADO DEL LIBRO EXCEL (.XLSX)
# -------------------------------------------------------------------
def generar_excel_formateado(anio, mes, weeks_data, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Distribucion_{mes}_{anio}"

    def apply_border(cell):
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    top_row_texts = {
        "Lunes": {"manana": "(Herrera HO)", "tarde": ""},
        "Martes": {"manana": "Aranda HO", "tarde": "Coronel"},
        "Mierc.": {"manana": "", "tarde": "Coronel"},
        "Jueves": {"manana": "", "tarde": ""},
        "Viernes": {"manana": "", "tarde": "(Coronel HO)"}
    }

    col = 2
    for day_name in ["Lunes", "Martes", "Mierc.", "Jueves", "Viernes"]:
        c1 = ws.cell(row=1, column=col, value=day_name)
        c2 = ws.cell(row=1, column=col+1)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        c1.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        c1.font = Font(bold=True)
        apply_border(c1)
        apply_border(c2)
        col += 2

    c1 = ws.cell(row=1, column=1, value=f"MES: {mes}/{anio}")
    c1.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    c1.font = Font(bold=True)
    apply_border(c1)

    row_idx = 2
    for w_num, days in weeks_data:
        ws.cell(row=row_idx, column=1, value="Lunes").font = Font(bold=True)
        apply_border(ws.cell(row=row_idx, column=1))
        
        col = 2
        for _, d_name in days:
            c1 = ws.cell(row=row_idx, column=col, value="mañana")
            c1.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            c1.font = Font(bold=True)
            apply_border(c1)
            
            c2 = ws.cell(row=row_idx, column=col+1, value="tarde")
            c2.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            c2.font = Font(color="FFFFFF", bold=True)
            apply_border(c2)
            col += 2
            
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value=f"Sem. {w_num}").font = Font(bold=True)
        apply_border(ws.cell(row=row_idx, column=1))
        
        col = 2
        for date, day_name_raw in days:
            day_name = normalizar_dia(day_name_raw)
            c1 = ws.cell(row=row_idx, column=col, value=top_row_texts.get(day_name, {}).get("manana", ""))
            c1.font = Font(color="7030A0", bold=True)
            apply_border(c1)
            
            c2 = ws.cell(row=row_idx, column=col+1, value=top_row_texts.get(day_name, {}).get("tarde", ""))
            c2.font = Font(color="7030A0", bold=True)
            apply_border(c2)
            col += 2
            
        row_idx += 1
        
        apply_border(ws.cell(row=row_idx, column=1))
        
        col = 2
        for date, day_name_raw in days:
            day_name = normalizar_dia(day_name_raw)
            if schedule[date]["feriado"]:
                c1 = ws.cell(row=row_idx, column=col, value="FERIADO")
                c2 = ws.cell(row=row_idx, column=col+1)
                ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col+1)
                c1.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
                c1.font = Font(color="FF0000", bold=True)
                apply_border(c1)
                apply_border(c2)
                col += 2
                continue
                
            manana_list = schedule[date]["manana"]
            tarde_list = schedule[date]["tarde"]
            early_person = schedule[date].get("early_person")
            
            manana_parts = []
            for p in manana_list:
                manana_parts.append(f"{p}*" if p == early_person else p)

            for p in schedule[date]["conae"]:
                manana_parts.append(f"{p} (CONAE)")
                    
            for p in schedule[date]["francos"]:
                manana_parts.append(f"{p} (FRANCO)")

            for emp, detail in schedule[date]["licencia"].items():
                lic_str = f"{emp} ({detail})" if detail else f"{emp} (LIC)"
                manana_parts.append(lic_str)
                
            manana_text = ", ".join(manana_parts)
            
            ho_group = ""
            if w_num % 2 != 0:
                if day_name == "Lunes": ho_group = "BHO"
                if day_name == "Viernes": ho_group = "AHO"
            else:
                if day_name == "Lunes": ho_group = "AHO"
                if day_name == "Viernes": ho_group = "BHO"
            
            # Si hay Home Office dinámicos adicionales hoy, agregarlos al texto HO
            ho_total = list(schedule[date]["ho"])
            if ho_group and ho_group not in ho_total:
                ho_total_str = f"{ho_group} (" + ", ".join(ho_total) + ")" if ho_total else ho_group
            else:
                ho_total_str = ", ".join(ho_total)

            if ho_total_str:
                manana_text += f"\n\n{ho_total_str}"
                
            c1 = ws.cell(row=row_idx, column=col, value=manana_text)
            apply_border(c1)
            
            tarde_text = "\n".join(tarde_list)
            c2 = ws.cell(row=row_idx, column=col+1, value=tarde_text)
            c2.font = Font(color="376092", bold=True)
            apply_border(c2)
            
            col += 2
            
        ws.row_dimensions[row_idx].height = 65
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="No hay franco los dias anteriores o posteriores al home office")
    ws.cell(row=row_idx+1, column=1, value="Herrera a criterio personal y/o acorde solicitud...")
    ws.cell(row=row_idx+2, column=1, value="Coronel, decidió la opción de 1 hora menos, no tiene franco")
    ws.cell(row=row_idx+3, column=1, value="Si vienen para pasarse CORE, solo son 5hs laborables...")
    note_cell = ws.cell(row=row_idx+4, column=1, value="*entran antes de 7 para monitores")
    note_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    note_cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 12
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 20

    return wb

# -------------------------------------------------------------------
# MANIPULADOR PRINCIPAL DE EVENTO DEL BOTÓN
# -------------------------------------------------------------------
async def ejecutar_generacion(event):
    box_estado = document.querySelector("#estado")
    box_estado.className = "status-box"
    
    try:
        mes = int(document.querySelector("#mes").value)
        anio = int(document.querySelector("#anio").value)
        
        box_estado.innerText = "🔍 Consultando feriados de Argentina en vivo..."
        feriados_api = await consultar_feriados(anio, mes)

        personal_json_str = document.querySelector("#personal_json").value
        personal_info = parse_personal_dinamico(personal_json_str)

        novedades_json_str = document.querySelector("#novedades_json").value
        conae_dict, licencias_dict, francos_solicitados_dict, mananas_solicitadas_dict, tardes_solicitadas_dict, ho_excepcional_dict, feriados_manuales = parse_novedades_json(novedades_json_str, anio, mes)

        feriados_totales = list(set(feriados_api + feriados_manuales))

        chks = document.querySelectorAll(".alfa-chk:checked")
        bloqueados_alfa = [chk.value for chk in chks]

        box_estado.innerText = f"⚙️ Generando matriz de turnos dinámicas..."

        weeks_data, schedule = resolver_planilla(
            anio, mes, feriados_totales, conae_dict, licencias_dict,
            francos_solicitados_dict, mananas_solicitadas_dict, tardes_solicitadas_dict, ho_excepcional_dict, bloqueados_alfa,
            personal_info
        )
        wb = generar_excel_formateado(anio, mes, weeks_data, schedule)

        filename = f"Planilla_Distribucion_{mes}_{anio}.xlsx"
        descargar_excel(wb, filename)

        box_estado.innerText = f"✅ ¡Planilla descargada con éxito! ({filename})"
        box_estado.className = "status-box success"

    except Exception as err:
        box_estado.innerText = f"❌ Error durante la generación: {str(err)}"
        box_estado.className = "status-box error"

btn_generar = document.querySelector("#btn-generar")
proxy_handler = create_proxy(ejecutar_generacion)
btn_generar.addEventListener("click", proxy_handler)
